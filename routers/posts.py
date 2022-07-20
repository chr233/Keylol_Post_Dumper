'''
# @Author       : Chr_
# @Date         : 2021-05-17 10:39:02
# @LastEditors  : Chr_
# @LastEditTime : 2022-07-20 13:37:29
# @Description  : 历史记录查询
'''


from fastapi.exceptions import HTTPException
from fastapi.responses import FileResponse
from fastapi.routing import APIRouter
from models.posts import Posts, Posts_In, Posts_Out
from models.response import (BaseResponse, PostDetailResponse, PostIdsResponse,
                             PostsListResponse)
from openpyxl import Workbook

router = APIRouter(prefix='/api',
                   tags=['帖子数据接口'],
                   responses={404: {'detail': 'page not found'}},)


@router.get('/test', response_model=BaseResponse)
async def api_test():
    return {'code': 666, 'msg': 'server is ready'}


@router.get('/posts', response_model=PostsListResponse)
async def get_post_list(page: int = 1, limit: int = 50):
    '''
    读取帖子列表
    '''
    if (page < 1 or limit < 10):
        return {'code': 422, 'msg': 'params error', 'data': None}

    offset = (page-1) * limit
    qs = Posts.all().offset(offset).limit(limit)

    post_list = await Posts_Out.from_queryset(qs)
    count = await Posts.all().count()

    return {'count': count, 'data': post_list}


@router.get('/posts/ids', response_model=PostIdsResponse)
async def get_post_list():
    '''
    读取帖子列表
    '''

    post_list = await Posts.all().only('tid')

    id_list = [x.tid for x in post_list]

    count = await Posts.all().count()

    return {'count': count, 'data': id_list}


@router.get('/post/{tid}', response_model=PostDetailResponse)
async def get_post_list(tid: str):
    '''
    读取帖子信息
    '''
    post = await Posts.filter(tid=tid).get_or_none()

    if not post:
        return {'code': 404, 'msg': 'not found', 'data': None}

    else:
        return {'code': 0, 'msg': 'ok', 'data': post}


@router.post('/post', response_model=BaseResponse)
async def create_or_update_post(new_post: Posts_In):
    '''
    新建帖子
    '''
    post = await Posts.filter(tid=new_post.tid).get_or_none()

    if not post:
        post = await Posts.create(**new_post.dict(exclude_unset=True))
    else:
        post.update_from_dict(new_post.dict(exclude_unset=True))
        await post.save()

    return {'code': 0, 'msg': 'ok'}


@router.delete('/post/{tid}', response_model=BaseResponse)
async def delete_post(tid: str):
    '''
    删除帖子
    '''
    post = await Posts.filter(tid=tid).get_or_none()

    if not post:
        raise HTTPException(status_code=404, detail='post not found')

    else:
        await post.delete()

    return {'code': 0, 'msg': 'ok'}


@router.delete('/posts', response_model=BaseResponse)
async def delete_all_post():
    '''
    删除全部帖子
    '''
    await Posts.all().delete()

    return {'code': 0, 'msg': 'ok'}


@router.get("/excel", response_class=FileResponse)
async def as_excel():
    '''
    导出excel
    '''

    wb = Workbook()
    ws = wb.active
    ws.title = '帖子列表'

    all_posts = await Posts.all()

    ws.cell(row=1, column=1).value = '帖子ID'
    ws.cell(row=1, column=2).value = '链接'
    ws.cell(row=1, column=3).value = '标题'
    ws.cell(row=1, column=4).value = '作者昵称'
    ws.cell(row=1, column=5).value = '作者ID'
    ws.cell(row=1, column=6).value = '作者链接'
    ws.cell(row=1, column=7).value = '发布时间'
    ws.cell(row=1, column=8).value = '帖子正文'
    ws.cell(row=1, column=9).value = '提及游戏ID'
    ws.cell(row=1, column=9).value = '提及游戏链接'

    i = 2

    for post in all_posts:
        ws.cell(row=i, column=1).value = post.tid
        ws.cell(row=i, column=2).value = post.post_url
        ws.cell(row=i, column=3).value = post.post_title
        ws.cell(row=i, column=4).value = post.author_nick
        ws.cell(row=i, column=5).value = post.author_uid
        ws.cell(
            row=i, column=6).value = f'https://keylol.com/suid-{post.author_uid}'
        ws.cell(row=i, column=7).value = post.post_date
        ws.cell(row=i, column=8).value = post.content
        ws.cell(row=i, column=9).value = post.game_list
        ws.cell(row=i, column=9).value = post.game_excel
        i += 1

    wb.save('output.xlsx')
    return FileResponse('output.xlsx')


@router.get("/bbcode", response_class=FileResponse)
async def as_excel():
    '''
    导出bbcode
    '''

    all_posts = await Posts.all()
    lines = [
        '[table]',
        '[tr]',
        '[td]提及游戏[/td]',
        '[td]作者昵称[/td]',
        '[td]作者ID[/td]',
        '[td]发布时间[/td]',
        '[td]原帖链接[/td]',
        '[td]帖子正文[/td]',
        '[/tr]',
        ]

    for post in all_posts:
        lines.append('')
        lines.append('[tr]')
        lines.append(f'[td]{post.game_bbcode or post.game_list}[/td]')
        lines.append(f'[td]{post.author_nick}[/td]')
        lines.append(f'[td]{post.author_uid}[/td]')
        lines.append(f'[td]{post.post_date}[/td]')
        lines.append(f'[td][url={post.post_url}]{post.post_title}[/url][/td]')
        lines.append(f'[td]{post.content}[/td]')
        lines.append('[/tr]')

    lines.append('[/table]')

    with open('output.txt', 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

    return FileResponse('output.txt')
